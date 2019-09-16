# -*- coding: utf-8 -*-
"""
Created on Thu Aug  1 01:16:08 2019

@author: Batuhan Ökmen
"""
 


def anaMenu():
    
    global menusecim
    
    print("""
  ═════════════════════════════════════
  1- Hesapla
  2- Çıkış
  ═════════════════════════════════════ """)   









  
 #tanımlama yaptım kolaylık için
def Hesapla():  
     
        import openpyxl
        kitap = openpyxl.load_workbook("gelir.xlsx")
        sheet=kitap.active
        #süt değerlerini giriyoruz tanımlama işlemi başladı
        sabah1=sheet['B68']
        sabah2=sheet['B69']
        sabah3=sheet['B70']
        sabah4=sheet['B71']
        sabah5=sheet['B72']
        sabah6=sheet['B73']
        sabah7=sheet['B74']
        
        aksam1=sheet['C68']
        aksam2=sheet['C69']
        aksam3=sheet['C70']
        aksam4=sheet['C71']
        aksam5=sheet['C72']
        aksam6=sheet['C73']
        aksam7=sheet['C74']
        
        SutSonuc1=sheet['D68']
        SutSonuc2=sheet['D69']
        SutSonuc3=sheet['D70']
        SutSonuc4=sheet['D71']
        SutSonuc5=sheet['D72']
        SutSonuc6=sheet['D73']
        SutSonuc7=sheet['D74']
        
        
        #       TAHMİN SÜT DEĞERLERİ
        
        BuGunS=sheet['B6']
        UcAyOnceS=sheet['B7']
        AltiAyOnceS=sheet['B8']
        UcAySonraS=sheet['B9']
        AltiAySonraS=sheet['B10']
        OnİkiAySonraS=sheet['B11']
        
        BuGunA=sheet['C6']
        UcAyOnceA=sheet['C7']
        AltiAyOnceA=sheet['C8']
        UcAySonraA=sheet['C9']
        AltiAySonraA=sheet['C10']
        OnİkiAySonraA=sheet['C11']
        
        BuGunTl=sheet['D6']
        UcAyOnceTl=sheet['D7']
        AltiAyOnceTl=sheet['D8']
        UcAySonraTl=sheet['D9']
        AltiAySonraTl=sheet['D10']
        OnİkiAySonraTl=sheet['D11']
        
        #       TAHMİN YEM ₺ DEĞERLERİ
        
        SutYemiBugun=sheet['B14']
        UcAyOnceSu=sheet['B15']
        AltiAyOnceSu=sheet['B16']
        OnİkiAyOnceSu=sheet['B17']
        UcAySonraSu=sheet['B18']
        AltiAySonraSu=sheet['B19']
        OnİkiAySonraSu=sheet['B20']
        
        BesiYemiBugun=sheet['C14']
        UcAyOnceB=sheet['C15']
        AltiAyOnceB=sheet['C16']
        OnİkiAyOnceB=sheet['C17']
        UcAySonraB=sheet['C18']
        AltiAySonraB=sheet['C19']
        OnİkiAySonraB=sheet['C20']
        
        BuzagiYemiBugun=sheet['D14']
        UcAyOnceBU=sheet['D15']
        AltiAyOnceBU=sheet['D16']
        OnİkiAyOnceBU=sheet['D15']
        UcAySonraBU=sheet['D17']
        AltiAySonraBU=sheet['D18']
        OnİkiAySonraBU=sheet['D19']
        
        #       TAHMİN YEM ADET DEĞERLERİ
        
        SutYemiBugunA=sheet['B23']
        UcAyOnceSuA=sheet['B24']
        AltiAyOnceSuA=sheet['B25']
        OnİkiAyOnceSuA=sheet['B26']
        UcAySonraSuA=sheet['B27']
        AltiAySonraSuA=sheet['B28']
        OnİkiAySonraSuA=sheet['B29']
        
        BesiYemiBugunA=sheet['C23']
        UcAyOnceBA=sheet['C24']
        AltiAyOnceBA=sheet['C25']
        OnİkiAyOnceBA=sheet['C26']
        UcAySonraBA=sheet['C27']
        AltiAySonraBA=sheet['C28']
        OnİkiAySonraBA=sheet['C29']
        
        BuzagiYemiBugunA=sheet['D23']
        UcAyOnceBUA=sheet['D24']
        AltiAyOnceBUA=sheet['D25']
        OnİkiAyOnceBUA=sheet['D26']
        UcAySonraBUA=sheet['D27']
        AltiAySonraBUA=sheet['D28']
        OnİkiAySonraBUA=sheet['D29']
        
        #       TAHMİN BAKIM DEĞERLERİ
        
        
        AltiAyOnceBAK=sheet['B33']
        OnİkiAyOnceBAK=sheet['B36']
        AltiAySonraBAK=sheet['B37']
        OnİkiAySonraBAK=sheet['B38']
        
        # KAZANILAN TL MİKTARI
        
        buaydakazanılantlS=sheet['B78']
        ucayönceaydakazanılantlS=sheet['B79']
        altıayönceaydakazanılantlS=sheet['B80']
        onikiayönceaydakazanılantlS=sheet['B81']
        ucaysonraaydakazananılacaktlS=sheet['B82']
        altıaysonraaydakazananılacaktlS=sheet['B83']
        onikiaysonraaydakazananılacaktlS=sheet['B84']
        
        buaydakazanılantlA=sheet['C78']
        ucayönceaydakazanılantlA=sheet['C79']
        altıayönceaydakazanılantlA=sheet['C80']
        onikiayönceaydakazanılantlA=sheet['C81']
        ucaysonraaydakazananılacaktlA=sheet['C82']
        altıaysonraaydakazananılacaktlA=sheet['C83']
        onikiaysonraaydakazananılacaktlA=sheet['C84']
        
        buaydakazanılantlT=sheet['D78']
        ucayönceaydakazanılantlT=sheet['D79']
        altıayönceaydakazanılantlT=sheet['D80']
        onikiayönceaydakazanılantlT=sheet['D81']
        ucaysonraaydakazananılacaktlT=sheet['D82']
        altıaysonraaydakazananılacaktlT=sheet['D83']
        onikiaysonraaydakazananılacaktlT=sheet['D84']


        
        
        
        
        
        
        
        
        
        

        
        
        #çalışma kontrolü
        print(BuGunTl.value,':',sabah4.value)

        kitap.close()
        
#yazalım bakalım float yapmadan ilerleme kararı aldım sonrasında basım ağrırsa ilerdeki batu dusunsun bu olayı artık see you :D
        SutSonuc1=(aksam1.value + sabah1.value)
        SutSonuc2=(sabah2.value + aksam2.value)
        SutSonuc3=(aksam3.value + sabah3.value)
        SutSonuc4=(aksam4.value + sabah4.value)
        SutSonuc5=(aksam5.value + sabah5.value)
        SutSonuc6=(aksam6.value + sabah6.value)
        SutSonuc7=(aksam7.value + sabah7.value)
        
        #yem giderleri
        #gün
        BuAyHarcananTlSut =(SutYemiBugunA.value * SutYemiBugun.value )
        BuAyHarcananTlBesi =(BesiYemiBugunA.value * BesiYemiBugun.value )
        BuAyHarcananTlBuza =(BuzagiYemiBugunA.value * BuzagiYemiBugun.value )
        BuAyHarcananTlToplam =(BuAyHarcananTlSut + BuAyHarcananTlBesi + BuAyHarcananTlBuza)
        
        #3 ay önce
        UcAyOnceHarcananTlSut =(UcAyOnceSuA.value * UcAyOnceSu.value )
        UcAyOnceHarcananTlBesi =(UcAyOnceBA.value * UcAyOnceB.value )
        UcAyOnceHarcananTlBuza =(UcAyOnceBUA.value * UcAyOnceBU.value )
        UcAyOnceHarcananTlToplam =(UcAyOnceHarcananTlSut + UcAyOnceHarcananTlBesi + UcAyOnceHarcananTlBuza)    

        #6 ay önce
        
        AltiAyOnceHarcananTlSut =(AltiAyOnceSuA.value * AltiAyOnceSu.value )
        AltiAyOnceHarcananTlBesi =(AltiAyOnceBA.value * AltiAyOnceB.value )
        AltiAyOnceHarcananTlBuza =(AltiAyOnceBUA.value * AltiAyOnceBU.value )
        AltiAyOnceHarcananTlToplam =(AltiAyOnceHarcananTlSut + AltiAyOnceHarcananTlBesi + AltiAyOnceHarcananTlBuza)
        
        # 12 ay önce
        
        OnİkiAyOnceHarcananTlSut =(OnİkiAyOnceSuA.value * OnİkiAyOnceSu.value )
        OnİkiAyOnceHarcananTlBesi =(OnİkiAyOnceBA.value * OnİkiAyOnceB.value )
        OnİkiAyOnceHarcananTlBuza =(OnİkiAyOnceBUA.value * OnİkiAyOnceBU.value )
        OnİkiAyOnceHarcananTlToplam =(OnİkiAyOnceHarcananTlSut + OnİkiAyOnceHarcananTlBesi + OnİkiAyOnceHarcananTlBuza)


        
        #3 ay sonra
        UcAySonraHarcananTlSut =(UcAySonraSuA.value * UcAySonraSu.value )
        UcAySonraHarcananTlBesi =(UcAySonraBA.value * UcAySonraB.value )
        UcAySonraHarcananTlBuza =(UcAySonraBUA.value * UcAySonraBU.value )
        UcAySonraHarcananTlToplam =(UcAySonraHarcananTlSut + UcAySonraHarcananTlBesi + UcAySonraHarcananTlBuza)    

        #6 ay sonra
        
        AltiAySonraHarcananTlSut =(AltiAySonraSuA.value * AltiAySonraSu.value )
        AltiAySonraHarcananTlBesi =(AltiAySonraBA.value * AltiAySonraB.value )
        AltiAySonraHarcananTlBuza =(AltiAySonraBUA.value * AltiAySonraBU.value )
        AltiAySonraHarcananTlToplam =(AltiAySonraHarcananTlSut + AltiAySonraHarcananTlBesi + AltiAySonraHarcananTlBuza)
        
        # 12 ay sonra
        
        OnİkiAySonraHarcananTlSut =(OnİkiAySonraSuA.value * OnİkiAySonraSu.value )
        OnİkiAySonraHarcananTlBesi =(OnİkiAySonraBA.value * OnİkiAySonraB.value )
        OnİkiAySonraHarcananTlBuza =(OnİkiAySonraBUA.value * OnİkiAySonraBU.value )
        OnİkiAySonraHarcananTlToplam =(OnİkiAySonraHarcananTlSut + OnİkiAySonraHarcananTlBesi + OnİkiAySonraHarcananTlBuza)

        # KAZANILAN TL MİKTARINI HESAPLIYORUM
        
        sabahToplamSut=(sabah1.value + sabah2.value + sabah3.value + sabah4.value + sabah5.value + sabah6.value + sabah7.value)
        aksamToplamSut=(aksam1.value + aksam2.value + aksam3.value + aksam4.value + aksam5.value + aksam6.value + aksam7.value)
        buaydakazanılantlS =(sabahToplamSut * 4 * float(BuGunTl.value))
        buaydakazanılantlA =(aksamToplamSut * 4 * float(BuGunTl.value))
        buaydakazanılantlT =( buaydakazanılantlS + buaydakazanılantlA )
        
        # 3 AY ÖNCE KAZANILAN TL
        
        UcAyOncekazanılantlS =(UcAyOnceS.value * 30 * float(UcAyOnceTl.value))
        UcAyOncekazanılantlA =(UcAyOnceA.value * 30 * float(UcAyOnceTl.value))
        UcAyOncekazanılantlT =( UcAyOncekazanılantlS + UcAyOncekazanılantlA )
        
        # 6 AY ÖNCE KAZANILAN TL
        
        AltiAyOncekazanılantlS =(AltiAyOnceS.value * 30 * float(AltiAyOnceTl.value))
        AltiAyOncekazanılantlA =(AltiAyOnceA.value * 30 * float(AltiAyOnceTl.value))
        AltiAyOncekazanılantlT =( AltiAyOncekazanılantlS + AltiAyOncekazanılantlA )
        
        # 3 AY SONRA KAZANILAN TL
        
        UcAySonrakazanılantlS =(UcAySonraS.value * 30 * float(UcAySonraTl.value))
        UcAySonrakazanılantlA =(UcAySonraA.value * 30 * float(UcAySonraTl.value))
        UcAySonrakazanılantlT =( UcAySonrakazanılantlS + UcAySonrakazanılantlA )
        
        # 6 AY SONRA KAZANILAN TL
        
        AltiAySonrakazanılantlS =(AltiAySonraS.value * 30 * float(AltiAySonraTl.value))
        AltiAySonrakazanılantlA =(AltiAySonraA.value * 30 * float(AltiAySonraTl.value))
        AltiAySonrakazanılantlT =( AltiAySonrakazanılantlS + AltiAySonrakazanılantlA )
        
        # 12 AY SONRA KAZANILAN TL
        
        OnİkiAySonrakazanılantlS =(OnİkiAySonraS.value * 30 * float(OnİkiAySonraTl.value))
        OnİkiAySonrakazanılantlA =(OnİkiAySonraA.value * 30 * float(OnİkiAySonraTl.value))
        OnİkiAySonrakazanılantlT =( OnİkiAySonrakazanılantlS + OnİkiAySonrakazanılantlA )

















        #kaydetme bu kısım çok fazla şey eklenecek 

        
        import openpyxl
        kitap = openpyxl.load_workbook("gelir.xlsx")
        sayfa = kitap.get_sheet_by_name("Sheet1")
        #süt sonuçlarını yazıyorum
        sayfa.cell(row=68,column=4,value=SutSonuc1)
        sayfa.cell(row=69,column=4,value=SutSonuc2)
        sayfa.cell(row=70,column=4,value=SutSonuc3)
        sayfa.cell(row=71,column=4,value=SutSonuc4)
        sayfa.cell(row=72,column=4,value=SutSonuc5)
        sayfa.cell(row=73,column=4,value=SutSonuc6)
        sayfa.cell(row=74,column=4,value=SutSonuc7)
        #yem giderlerini yazıyorum
        sayfa.cell(row=68,column=7,value=BuAyHarcananTlSut)
        sayfa.cell(row=68,column=8,value=BuAyHarcananTlBesi)
        sayfa.cell(row=68,column=9,value=BuAyHarcananTlBuza)
        sayfa.cell(row=68,column=10,value=BuAyHarcananTlToplam)
        #3 ay once
        sayfa.cell(row=69,column=7,value=UcAyOnceHarcananTlSut)
        sayfa.cell(row=69,column=8,value=UcAyOnceHarcananTlBesi)
        sayfa.cell(row=69,column=9,value=UcAyOnceHarcananTlBuza)
        sayfa.cell(row=69,column=10,value=UcAyOnceHarcananTlToplam)
        #6 ay once
        sayfa.cell(row=70,column=7,value=AltiAyOnceHarcananTlSut)
        sayfa.cell(row=70,column=8,value=AltiAyOnceHarcananTlBesi)
        sayfa.cell(row=70,column=9,value=AltiAyOnceHarcananTlBuza)
        sayfa.cell(row=70,column=10,value=AltiAyOnceHarcananTlToplam)
        #12 ay once
        sayfa.cell(row=71,column=7,value=OnİkiAyOnceHarcananTlSut)
        sayfa.cell(row=71,column=8,value=OnİkiAyOnceHarcananTlBesi)
        sayfa.cell(row=71,column=9,value=OnİkiAyOnceHarcananTlBuza)
        sayfa.cell(row=71,column=10,value=OnİkiAyOnceHarcananTlToplam)
        #3 ay sonra
        sayfa.cell(row=72,column=7,value=UcAySonraHarcananTlSut)
        sayfa.cell(row=72,column=8,value=UcAySonraHarcananTlBesi)
        sayfa.cell(row=72,column=9,value=UcAySonraHarcananTlBuza)
        sayfa.cell(row=72,column=10,value=UcAySonraHarcananTlToplam)
        #6 ay sonra
        sayfa.cell(row=73,column=7,value=AltiAySonraHarcananTlSut)
        sayfa.cell(row=73,column=8,value=AltiAySonraHarcananTlBesi)
        sayfa.cell(row=73,column=9,value=AltiAySonraHarcananTlBuza)
        sayfa.cell(row=73,column=10,value=AltiAySonraHarcananTlToplam)
        #12 ay sonra
        sayfa.cell(row=74,column=7,value=OnİkiAySonraHarcananTlSut)
        sayfa.cell(row=74,column=8,value=OnİkiAySonraHarcananTlBesi)
        sayfa.cell(row=74,column=9,value=OnİkiAySonraHarcananTlBuza)
        sayfa.cell(row=74,column=10,value=OnİkiAySonraHarcananTlToplam)
        
        		#SABAH KAZANILAN TL 
        sayfa.cell(row=78,column=2,value=buaydakazanılantlS)
        sayfa.cell(row=79,column=2,value=UcAyOncekazanılantlS)
        sayfa.cell(row=80,column=2,value=AltiAyOncekazanılantlS)
        sayfa.cell(row=82,column=2,value=UcAySonrakazanılantlS)
        sayfa.cell(row=83,column=2,value=AltiAySonrakazanılantlS)
        sayfa.cell(row=84,column=2,value=OnİkiAySonrakazanılantlS)
		
		#AKŞAM KAZANILAN TL 
        sayfa.cell(row=78,column=3,value=buaydakazanılantlA)
        sayfa.cell(row=79,column=3,value=UcAyOncekazanılantlA)
        sayfa.cell(row=80,column=3,value=AltiAyOncekazanılantlA)
        sayfa.cell(row=82,column=3,value=UcAySonrakazanılantlA)
        sayfa.cell(row=83,column=3,value=OnİkiAySonrakazanılantlA)
        sayfa.cell(row=84,column=3,value=OnİkiAySonrakazanılantlS)
		
		#TOPLAM KAZANILAN TL 
        sayfa.cell(row=78,column=4,value=buaydakazanılantlT)
        sayfa.cell(row=79,column=4,value=UcAyOncekazanılantlT)
        sayfa.cell(row=80,column=4,value=AltiAyOncekazanılantlT)
        sayfa.cell(row=82,column=4,value=UcAySonrakazanılantlT)
        sayfa.cell(row=83,column=4,value=AltiAySonrakazanılantlT)
        sayfa.cell(row=84,column=4,value=OnİkiAySonrakazanılantlT)

        
        kitap.save("gelir.xlsx")
        kitap.close()


        print("ana menüye gitmek istermisiniz evet/hayır")
        secim= input("lütfen seçiminizi yapınız: ")
        if secim == ("evet") :
            print("Ana sayfaya gidiyorsunuz")
            anaMenu()
        else:
            print("tamam ozaman ben kaçtım")
            input("bir tuşa bas ")



while True:
    anaMenu()
    secim= int(input("lütfen seçiminizi yapınız: "))
    print("İŞLEM TAMAMLANIYOR LÜTFEN BEKLEYİNİZ")
    if secim == 1:
        print("İŞLEM TAMAMLANIYOR LÜTFEN BEKLEYİNİZ")
        Hesapla()
    elif secim == 5:
        pass
    elif secim == 2:
        print("çıkış yapılıyor")
        break
    else:
        print("bence seçimin yanlış birdaha dene...")












